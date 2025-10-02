'''
This script will read in the raw tracking data into a npy format
This format is faster and more convienient to read in downstream processing
The rotation of each participant is also converted into rotation matricies
Since we may not know the Euler rotation order, this script will output version for all of them
'''

from openpyxl import load_workbook
import numpy as np
from tqdm import tqdm
import scipy
import sys
import os
import pandas as pd
import os
from multiprocessing import Pool

all_rotation_orders = [
    'XYZ',
    'XZY',
    'YZX',
    'YXZ',
    'ZYX',
    'ZXY',
]



# for rotation_order in all_rotation_orders:
def process_data(rotation_order):
    sequence_code = sys.argv[1]
    print(f'Processing {os.path.abspath(f"raw_data_exp1/{sequence_code}/{sequence_code}.xlsx")}')
    df = pd.read_excel(f'raw_data_exp1/{sequence_code}/{sequence_code}.xlsx')

    print(f'Processing {os.path.abspath(f"raw_data_exp1/{sequence_code}/{sequence_code}.xlsx")}')

    # Using pandas, column headers are directly available as df.columns
    column_map = {col_name: idx for idx, col_name in enumerate(df.columns)}

    data_out = {
        'T1':{
            'eye_2d':[],
            'translation':[],
            'rot_mat':[],
            'rot_euler':[],
            'left_gaze_dir':[],
            'right_gaze_dir':[],
            'left_gaze_origin':[],
            'right_gaze_origin':[],
            'gaze_3d':[],
        },
        'T2':{
            'eye_2d':[],
            'translation':[],
            'rot_mat':[],
            'rot_euler':[],
            'left_gaze_dir':[],
            'right_gaze_dir':[],
            'left_gaze_origin':[],
            'right_gaze_origin':[],
            'gaze_3d':[],
        }
    }
    for row in tqdm(range(2,len(df))):
        # DEBUG
        # x = df.iloc[row][column_map['T1.leftGazeOrigin.X']
        # y = df.iloc[row][column_map['T1.leftGazeOrigin.Y']
        # z = df.iloc[row][column_map['T1.leftGazeOrigin.Z']
        # left_o = np.array([x,y,z]).astype(np.float32)
        # x = df.iloc[row][column_map['T1.rightGazeOrigin.X']
        # y = df.iloc[row][column_map['T1.rightGazeOrigin.Y']
        # z = df.iloc[row][column_map['T1.rightGazeOrigin.Z']
        # right_o = np.array([x,y,z]).astype(np.float32)
        # print(left_o,right_o)

        # ========================================

        # extract T1 2d eye gaze
        x = df.iloc[row][column_map['T1.gaze2D.X']]
        y = df.iloc[row][column_map['T1.gaze2D.Y']]
        data_out['T1']['eye_2d'].append(np.array([x,y]).astype(np.float32))
        #extract T1 3d gaze
        x = df.iloc[row][column_map['T1.gaze3D.X']]
        y = df.iloc[row][column_map['T1.gaze3D.Y']]
        z = df.iloc[row][column_map['T1.gaze3D.Z']]
        data_out['T1']['gaze_3d'].append(np.array([x,y,z]).astype(np.float32))
        # extract T1 position
        x = df.iloc[row][column_map['T1.TX']]
        y = df.iloc[row][column_map['T1.TY']]
        z = df.iloc[row][column_map['T1.TZ']]
        pos = np.array([x,y,z])
        data_out['T1']['translation'].append(pos.astype(np.float32))

        # extract T1 euler rotation
        x = df.iloc[row][column_map['T1.RX']]
        y = df.iloc[row][column_map['T1.RY']]
        z = df.iloc[row][column_map['T1.RZ']]
        rot = scipy.spatial.transform.Rotation.from_euler(rotation_order,[x,y,z],degrees=True)
        data_out['T1']['rot_mat'].append(rot.as_matrix().astype(np.float32))
        data_out['T1']['rot_euler'].append(np.array([x,y,z]).astype(np.float32))

        # extract T1 gaze origin
        x = df.iloc[row][column_map['T1.leftGazeOrigin.X']]
        y = df.iloc[row][column_map['T1.leftGazeOrigin.Y']]
        z = df.iloc[row][column_map['T1.leftGazeOrigin.Z']]
        gaze_origin_vector = np.array([x,y,z])
        gaze_origin_vector = rot.as_matrix() @ gaze_origin_vector + pos
        data_out['T1']['left_gaze_origin'].append(gaze_origin_vector.astype(np.float32))

        x = df.iloc[row][column_map['T1.rightGazeOrigin.X']]
        y = df.iloc[row][column_map['T1.rightGazeOrigin.Y']]
        z = df.iloc[row][column_map['T1.rightGazeOrigin.Z']]
        gaze_origin_vector = np.array([x,y,z])
        gaze_origin_vector = rot.as_matrix() @ gaze_origin_vector + pos
        data_out['T1']['right_gaze_origin'].append(gaze_origin_vector.astype(np.float32))

        # extract T1 gaze dir
        x = df.iloc[row][column_map['T1.leftGazeDirection.X']]
        y = df.iloc[row][column_map['T1.leftGazeDirection.Y']]
        z = df.iloc[row][column_map['T1.leftGazeDirection.Z']]
        gaze_dir_vector = np.array([x,y,z])
        gaze_dir_vector = rot.as_matrix() @ gaze_dir_vector
        data_out['T1']['left_gaze_dir'].append(gaze_dir_vector.astype(np.float32))

        x = df.iloc[row][column_map['T1.rightGazeDirection.X']]
        y = df.iloc[row][column_map['T1.rightGazeDirection.Y']]
        z = df.iloc[row][column_map['T1.rightGazeDirection.Z']]
        gaze_dir_vector = np.array([x,y,z])
        gaze_dir_vector = rot.as_matrix() @ gaze_dir_vector
        data_out['T1']['right_gaze_dir'].append(gaze_dir_vector.astype(np.float32))

        # extract T2 2d eye gaze
        x = df.iloc[row][column_map['T2.gaze2D.X']]
        y = df.iloc[row][column_map['T2.gaze2D.Y']]
        data_out['T2']['eye_2d'].append(np.array([x,y]).astype(np.float32))

       #extract T2 3d gaze
        x = df.iloc[row][column_map['T2.gaze3D.X']]
        y = df.iloc[row][column_map['T2.gaze3D.Y']]
        z = df.iloc[row][column_map['T2.gaze3D.Z']]
        data_out['T2']['gaze_3d'].append(np.array([x,y,z]).astype(np.float32))


        # extract T2 position
        x = df.iloc[row][column_map['T2.TX']]
        y = df.iloc[row][column_map['T2.TY']]
        z = df.iloc[row][column_map['T2.TZ']]
        pos = np.array([x,y,z])
        data_out['T2']['translation'].append(pos.astype(np.float32))

        # extract T2 euler rotation
        x = df.iloc[row][column_map['T2.RX']]
        y = df.iloc[row][column_map['T2.RY']]
        z = df.iloc[row][column_map['T2.RZ']]
        rot = scipy.spatial.transform.Rotation.from_euler(rotation_order,[x,y,z],degrees=True)
        data_out['T2']['rot_mat'].append(rot.as_matrix().astype(np.float32))
        data_out['T2']['rot_euler'].append(np.array([x,y,z]).astype(np.float32))

        # extract T2 gaze origin
        x = df.iloc[row][column_map['T2.leftGazeOrigin.X']]
        y = df.iloc[row][column_map['T2.leftGazeOrigin.Y']]
        z = df.iloc[row][column_map['T2.leftGazeOrigin.Z']]
        gaze_origin_vector = np.array([x,y,z])
        gaze_origin_vector = rot.as_matrix() @ gaze_origin_vector + pos
        data_out['T2']['left_gaze_origin'].append(gaze_origin_vector.astype(np.float32))

        x = df.iloc[row][column_map['T2.rightGazeOrigin.X']]
        y = df.iloc[row][column_map['T2.rightGazeOrigin.Y']]
        z = df.iloc[row][column_map['T2.rightGazeOrigin.Z']]
        gaze_origin_vector = np.array([x,y,z])
        gaze_origin_vector = rot.as_matrix() @ gaze_origin_vector + pos
        data_out['T2']['right_gaze_origin'].append(gaze_origin_vector.astype(np.float32))
        
        # extract T2 gaze dir
        x = df.iloc[row][column_map['T2.leftGazeDirection.X']]
        y = df.iloc[row][column_map['T2.leftGazeDirection.Y']]
        z = df.iloc[row][column_map['T2.leftGazeDirection.Z']]
        gaze_dir_vector = np.array([x,y,z])
        gaze_dir_vector = rot.as_matrix() @ gaze_dir_vector
        data_out['T2']['left_gaze_dir'].append(gaze_dir_vector.astype(np.float32))

        x = df.iloc[row][column_map['T2.rightGazeDirection.X']]
        y = df.iloc[row][column_map['T2.rightGazeDirection.Y']]
        z = df.iloc[row][column_map['T2.rightGazeDirection.Z']]
        gaze_dir_vector = np.array([x,y,z])
        gaze_dir_vector = rot.as_matrix() @ gaze_dir_vector
        data_out['T2']['right_gaze_dir'].append(gaze_dir_vector.astype(np.float32))

    os.makedirs(f'preprocessed/{sequence_code}',exist_ok=True)
    np.save(f'preprocessed/{sequence_code}/extracted_data_{rotation_order}.npy',data_out)


if __name__ == "__main__":
    # print(f'Processing {os.path.abspath(f"raw_data_exp1/{sequence_code}/{sequence_code}.xlsx")}')
    # wb = load_workbook(filename=f'raw_data_exp1/{sequence_code}/{sequence_code}.xlsx')
    # print(f'Processing {os.path.abspath(f"raw_data_exp1/{sequence_code}/{sequence_code}.xlsx")}')
    # sheet = wb.active

    # column_map = {}
    # for col in range(1, df.max_column+1):
    #     cell = df.cell(row=1,column=col)
    #     column_map[cell.value] = col
    # # print(column_map.keys()) # debug to see all columns

    with Pool(8) as p:
        p.map(process_data,all_rotation_orders)
